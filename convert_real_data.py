"""
convert_real_data.py
Converts the two real Zycus client project exports (Smartsheet-style Gantt
plans) into the task_list.csv + weekly_update.csv shape that
agent/ingest.py expects.

Refactored to take explicit paths (rather than hardcoded ones) so it can be
called two ways:
  1. As a standalone script (uses the bundled files in data/raw_xlsx/ by
     default): `python3 convert_real_data.py`
  2. As a library, e.g. by dashboard/app.py, which re-runs a single
     project's conversion on demand whenever it notices the source .xlsx
     file has been edited more recently than the CSVs derived from it.

Milestone detection: both source files are hierarchical outlines (top-level
project row, then phase rows, then task rows). Phase-level rows are treated
as milestones:
  - UniSan file: rows where Ancestors == 1 (phase-level in that hierarchy)
  - Titan/Outokumpu file: rows where Level == 1 (phase-level in that hierarchy)
Top-level project/phase ROLLUP rows (Ancestors==0 / Level==0) are excluded
from the task list entirely since they're aggregate summaries, not tasks.
"""
import openpyxl
import csv
import os
import re

THIS_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_UNISAN_XLSX = os.path.join(THIS_DIR, "data", "raw_xlsx", "Project_Plan_B.xlsx")
DEFAULT_TITAN_XLSX = os.path.join(THIS_DIR, "data", "raw_xlsx", "S2P_Project__1_.xlsx")
DEFAULT_OUT_DIR = os.path.join(THIS_DIR, "data", "project_plans")


def fmt_date(d):
    if d is None:
        return ""
    if isinstance(d, str):
        return d
    return d.strftime("%Y-%m-%d")


def write_tasks_csv(path, rows):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["task_name", "planned_end", "actual_end", "percent_complete", "milestone", "owner"])
        for r in rows:
            w.writerow(r)


def write_update_csv(path, week_ending, stakeholder_notes, pm_comments, open_blockers):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["week_ending", "total_budget", "budget_planned_to_date", "budget_actual_to_date",
                     "open_blockers", "stakeholder_notes", "pm_comments"])
        w.writerow([week_ending, "", "", "", open_blockers, stakeholder_notes, pm_comments])


# ---------------------------------------------------------------------------
# UniSan
# ---------------------------------------------------------------------------
def convert_unisan(xlsx_path=DEFAULT_UNISAN_XLSX, out_dir=DEFAULT_OUT_DIR):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Project Plan"]
    rows = []
    for r in range(2, ws.max_row + 1):
        ancestors = ws.cell(row=r, column=6).value  # Ancestors
        if ancestors == 0:
            continue  # skip whole-project rollup row
        task_name = ws.cell(row=r, column=12).value  # Task Name
        pct = ws.cell(row=r, column=14).value  # % Complete (0-1 fraction)
        end = ws.cell(row=r, column=16).value  # End Date
        owner = ws.cell(row=r, column=32).value  # Assigned To
        if not task_name:
            continue
        pct100 = round((pct or 0) * 100, 1)
        milestone = "Y" if ancestors == 1 else "N"
        rows.append([task_name, fmt_date(end), "", pct100, milestone, owner or ""])

    os.makedirs(out_dir, exist_ok=True)
    write_tasks_csv(os.path.join(out_dir, "unisan_tasks_current.csv"), rows)

    ws_sum = wb["Summary"]
    summary = {ws_sum.cell(row=r, column=1).value: ws_sum.cell(row=r, column=2).value for r in range(1, 21)}
    as_of = summary.get("Today's Date")
    stakeholder_notes = (
        f"Source project plan summary (as of {fmt_date(as_of)}): "
        f"schedule health rated {summary.get('Schedule Health')}, "
        f"overall at-risk rating {summary.get('At Risk')}, "
        f"project stage '{summary.get('Project Stage')}', "
        f"{round((summary.get('% Complete') or 0) * 100)}% complete overall. "
        f"No stakeholder/PM free-text commentary was present in the source export."
    )
    write_update_csv(os.path.join(out_dir, "unisan_current_update.csv"), fmt_date(as_of), stakeholder_notes, "", "")
    return fmt_date(as_of), len(rows)


# ---------------------------------------------------------------------------
# Titan / Outokumpu
# ---------------------------------------------------------------------------
def convert_titan(xlsx_path=DEFAULT_TITAN_XLSX, out_dir=DEFAULT_OUT_DIR):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Outokumpu- S2P Project"]
    rows = []
    for r in range(2, ws.max_row + 1):
        level = ws.cell(row=r, column=5).value  # Level
        if level == 0:
            continue  # skip whole-project / phase rollup rows
        task_name = ws.cell(row=r, column=9).value  # Task Name
        pct = ws.cell(row=r, column=13).value  # % Complete (0-1 fraction, may be None)
        end = ws.cell(row=r, column=12).value  # End Date
        owner = ws.cell(row=r, column=19).value or ws.cell(row=r, column=30).value  # Owner, fallback Assigned To
        if not task_name:
            continue
        pct100 = round((pct or 0) * 100, 1)
        milestone = "Y" if level == 1 else "N"
        rows.append([task_name, fmt_date(end), "", pct100, milestone, owner or ""])

    os.makedirs(out_dir, exist_ok=True)
    write_tasks_csv(os.path.join(out_dir, "titan_tasks_current.csv"), rows)

    ws_sum = wb["Summary"]
    summary = {ws_sum.cell(row=r, column=1).value: ws_sum.cell(row=r, column=2).value for r in range(1, 21)}
    as_of = summary.get("Today's Date")

    ws_c = wb["Comments"]
    comment_notes = []
    for r in range(1, ws_c.max_row + 1):
        text = ws_c.cell(row=r, column=2).value
        author = ws_c.cell(row=r, column=3).value
        ts = ws_c.cell(row=r, column=4).value
        if text:
            comment_notes.append(f"{text.strip()} ({author}, {ts})" if author else text.strip())

    stakeholder_notes = (
        f"Source project plan summary (as of {fmt_date(as_of)}): "
        f"schedule health rated {summary.get('Schedule Health')}, "
        f"overall at-risk rating {summary.get('At Risk')}, "
        f"project stage '{summary.get('Project Stage')}', "
        f"{round((summary.get('% Complete') or 0) * 100)}% complete overall. "
        + (" ".join(comment_notes) if comment_notes else "No additional stakeholder commentary logged.")
    )
    # Only pull comments that actually describe a pending/blocking issue into
    # open_blockers (rather than every comment, which would mislabel routine
    # status updates like "we covered all sessions" as blockers).
    BLOCKER_KEYWORDS = re.compile(
        r"pending|impact|delay|gap|remain|blocked|waiting|issue|risk|need meeting|due to",
        re.IGNORECASE,
    )
    blocker_notes = [c for c in comment_notes if BLOCKER_KEYWORDS.search(c)]
    open_blockers = ";".join(blocker_notes[:3])
    write_update_csv(os.path.join(out_dir, "titan_current_update.csv"), fmt_date(as_of), stakeholder_notes, "", open_blockers)
    return fmt_date(as_of), len(rows)


PROJECTS = {
    "unisan": {"name": "UniSan S2P Implementation", "xlsx": DEFAULT_UNISAN_XLSX, "convert_fn": convert_unisan},
    "titan": {"name": "Titan (Outokumpu) S2P Implementation", "xlsx": DEFAULT_TITAN_XLSX, "convert_fn": convert_titan},
}


if __name__ == "__main__":
    as_of_u, n_u = convert_unisan()
    as_of_t, n_t = convert_titan()
    print(f"UniSan: {n_u} tasks, as-of {as_of_u}")
    print(f"Titan/Outokumpu: {n_t} tasks, as-of {as_of_t}")
